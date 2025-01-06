import PIL.Image
import os
import google.generativeai as genai
import openpyxl
import prompts
import time
import google.api_core.exceptions
import numpy as np

genai.configure(api_key="AIzaSyAcbaTLP7YJlAU4_OOfc2NVvAwVTkL6Yew")
folder_path = "images"
images = [
    os.path.join(folder_path, f)
    for f in os.listdir(folder_path)
    if f.lower().endswith((".jpg", ".jpeg", ".png"))
]

uploaded_files = [genai.upload_file(image_path) for image_path in images]

model = genai.GenerativeModel(model_name="gemini-1.5-flash")

wb = openpyxl.Workbook()
ws = wb.active

row = ws.max_row + 1
header = ["file name"]
for emotion in prompts.emotions:
    header.extend([emotion, None])
ws.append(header)

for idx, emotion in enumerate(prompts.emotions):
    col_start = 2 + idx * 2
    col_end = col_start + 1
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)

results_dict = {image_path: [] for image_path in images}

total_iterations = 50 * len(prompts.prompts)
current_iteration = 0



for _ in range(50):
    for prompt in prompts.prompts:
        while True:
            try:
                response = model.generate_content([prompt] + uploaded_files)
                print(response.text)
                scores = response.text.strip().split()
                print(scores)
                for i, image_path in enumerate(images):
                    results_dict[image_path].append(float(scores[i]))
                break
            except google.api_core.exceptions.ResourceExhausted:
                print("Quota exhausted. Retrying after a delay...")
                time.sleep(30)
            except google.api_core.exceptions.InvalidArgument as e:
                print(f"Invalid argument: {e}")
                break
        current_iteration += 1
        progress = (current_iteration / total_iterations) * 100
        print(f"Progress: {progress:.2f}%")

for image_path, scores in results_dict.items():
    avg_scores = np.mean(scores)
    std_dev_scores = np.std(scores)
    ws.append([os.path.basename(image_path), avg_scores, std_dev_scores])

wb.save("wyniki.xlsx")